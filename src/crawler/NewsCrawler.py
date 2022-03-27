from selenium import webdriver as wd
from selenium.webdriver.common.by import By
from bs4 import BeautifulSoup
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

import os
import re
import requests
import json
from openpyxl import load_workbook
import pandas as pd
import urllib
import time
from datetime import datetime
from tqdm import tqdm
from TR import TextRank

import math
from konlpy.tag import Okt
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.cluster import KMeans
from sklearn.metrics.pairwise import cosine_similarity

def set_chrome_driver() :
    chrome_options = wd.ChromeOptions()
    driver = wd.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
    return driver

def to_Excel(save_path, ranking_type, data_list) :
    # 엑셀 파일 생성
    if ranking_type == 'popular' :
        article_df = pd.DataFrame({'date' : data_list[0],
                                                   'press' : data_list[1],
                                                   'ranking' : data_list[2],
                                                   'views' : data_list[3],
                                                   'title' : data_list[4],
                                                   'link' : data_list[5],
                                                   'content' :data_list[6],
                                                   'keywords' : data_list[7]})
                                                    
    elif ranking_type == 'comment' :
        article_df = pd.DataFrame({"date" : data_list[0],
                                                   "press" : data_list[1],
                                                   "ranking" : data_list[2],
                                                   "comments" : data_list[3],
                                                   "title" : data_list[4],
                                                   "link" : data_list[5],
                                                   "content" :data_list[6],
                                                   "keywords" : data_list[7]})

    elif ranking_type == 'section' :
        article_df = pd.DataFrame({'date' : data_list[0],
                                                   'press' : data_list[1],
                                                   'section' : data_list[2],
                                                   'ranking' : data_list[3],
                                                   'views' : data_list[4],
                                                   'title' : data_list[5],
                                                   'link' : data_list[6],
                                                   'content' :data_list[7],
                                                   'keywords' : data_list[8]})
        
    elif ranking_type == 'keyword' :
        article_df = pd.DataFrame({"date" : data_list[0],
                                                   "press" : data_list[1],
                                                   "section" : data_list[2],
                                                   "reactions" : data_list[5],
                                                   "title" : data_list[3],
                                                   "link" : data_list[4],
                                                   "content" :data_list[6],
                                                   "keywords" : data_list[7]})

    print(f"extract article num : {len(article_df)}")
    article_df.to_excel(f"{save_path}.xlsx", index=False)

def to_Json(save_path, ranking_type, data_list) :

    data_dict = {}
    for idx in range (0, len(data_list[0])) :
        if ranking_type == 'popular' :
            data_dict[idx+1] = {'date' : data_list[0][idx], 'press' : data_list[1][idx], 'ranking' : data_list[2][idx], 'views' : data_list[3][idx], 'title' : data_list[4][idx], 'link' : data_list[5][idx], 'content' : data_list[6][idx], 'keyword' : data_list[7][idx]}
        elif ranking_type == 'comment' :
            data_dict[idx+1] = {'date' : data_list[0][idx], 'press' : data_list[1][idx],  'ranking' : data_list[2][idx], 'comments' : data_list[3][idx], 'title' : data_list[4][idx], 'link' : data_list[5][idx], 'content' : data_list[6][idx], 'keyword' : data_list[7][idx]}
        elif ranking_type == 'section' :
            data_dict[idx+1] = {'date' : data_list[0][idx], 'press' : data_list[1][idx],  'section' : data_list[2][idx], 'ranking' : data_list[3][idx], 'views' : data_list[4][idx], 'title' : data_list[5][idx], 'link' : data_list[6][idx], 'content' : data_list[7][idx], 'keyword' : data_list[8][idx]}
        elif ranking_type == 'keyword' :
            data_dict[idx+1] = {'date' : data_list[0][idx], 'press' : data_list[1][idx], 'section' : data_list[2][idx], 'reactions' : data_list[5][idx], 'title' : data_list[3][idx], 'link' : data_list[4][idx], 'content' : data_list[6][idx], 'keyword' : data_list[7][idx]}


    print(f"extract article num : {len(data_list[0])}")
    j = json.dumps(data_dict, ensure_ascii=False, indent="\t") 
    with open(f"{save_path}.json", 'w', encoding="utf-8") as f:
        f.write(j)


#########################
# 언론사별 랭킹뉴스 검색 #
#########################


def get_rankingNews(save_path, target_date, ranking_type) :

    crawl_date = f"{target_date[:4]}.{target_date[4:6]}.{target_date[6:]}"
    
    date_list, press_list, ranking_list, num_list, title_list, link_list, content_list, keyword_list = [], [], [], [], [], [], [], []
    if ranking_type == 'section' :
        section = input("섹션 (1. 정치, 2. 경제, 3. 사회, 4. 생활, 5. 세계, 6. IT) : ")

        sections = {'1' : '정치', '2' : '경제', '3' : '사회', '4' : '생활', '5' : '세계', '6' : 'IT'}
        section = sections[section]

        date_list, press_list, ranking_list, num_list, title_list, link_list = get_rankingNews_infos(crawl_date=crawl_date, ranking_type=ranking_type, section=section, is_section=True)

        section_list = []
        for i in range(len(date_list)) :
            section_list.append(section)

        data_list = [date_list, press_list, section_list, ranking_list, num_list, title_list, link_list]
        
    else :
        date_list, press_list, ranking_list, num_list, title_list, link_list = get_rankingNews_infos(crawl_date=crawl_date, ranking_type=ranking_type)
        data_list = [date_list, press_list, ranking_list, num_list, title_list, link_list]

    content_list, keyword_list = get_article_contents(link_list)
    data_list.append(content_list)
    data_list.append(keyword_list)

    #to_Excel(save_path, ranking_type, data_list)
    #to_Json(save_path, ranking_type, data_list)
    clustering(save_path, data_list, ranking_type)


def get_rankingNews_infos(crawl_date, ranking_type, section=None, is_section=False) :
    
    headers = {
        'referer' : 'https://www.naver.com/',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/99.0.4844.51 Safari/537.36'
        }

    date_list, press_list, ranking_list, num_list, title_list, link_list = [], [], [], [], [], []


    press_id = {'경향신문' : '032', '국민일보' : '005', '동아일보' : '020', '문화일보' : '021', '서울신문' : '081', '세계일보' : '022'}
    #press_id = {'경향신문' : '032', '국민일보' : '005', '동아일보' : '020', '문화일보' : '021', '서울신문' : '081', '세계일보' : '022',
                #'조선일보' : '023', '중앙일보' : '025', '한겨레' : '028', '한국일보' : '469', 'JTBC' : '437', 'KBS' : '056', 'MBC' : '214', 'MBN' : '057',
               #'SBS' : '055', 'SBS Biz' :  '374', 'TV 조선' : '448', 'YTN' : '052', '뉴스1' : '421', '뉴시스' : '003', '연합뉴스' : '001', '연합뉴스TV' : '422', '채널A' : '449', '한국경제TV' : '215',
                #'매일경제' : '009', '서울경제' : '011', '아시아경제' : '277', '한국경제 ' : '015', '헤럴드경제' : '016', '머니투데이' : '008', '이데일리' : '018', '조선비즈' : '366',
                #'매일신문' : '088', '강원일보' : '087', '부산일보' : '082'}

    for pid in list(press_id.values()) :
        url = f'https://media.naver.com/press/{pid}/ranking?type={ranking_type}&date={target_date}'

        res = requests.get(url, headers=headers)
        soup = BeautifulSoup(res.text, 'lxml')

        press = soup.select_one('div.press_hd_main_info > div > h3 > a').text
        press = re.sub('\xa0|\t|\r|\n|', '', press)
            
        if ranking_type == 'section' :

            child_nums = {'정치' : '3', '경제' : '4', '사회' : '5', '생활' : '6', '세계' : '7', 'IT' : '8'}
            child_num = child_nums[section]
            
            try :
                press_ranking_box = soup.select_one(f'div.press_ranking_home > div:nth-child({child_num})')
                rankingnews_lists = press_ranking_box.select('ul > li')

                for li in rankingnews_lists :
                    ranking = li.select_one('a > em').text
                    ranking_list.append(ranking)

                    title = li.select_one('.list_content > strong').text
                    title_list.append(title)
            
                    try :
                        cnt = li.select_one('.list_content > span').text
                        cnt = re.sub("조회수|,", "", cnt)
                        cnt = cnt.strip()
                        num_list.append(cnt)
                    except :
                        num_list.append('미제공')
                    
                    link = li.select_one('a')["href"]
                    link_list.append(link)

                    date_list.append(crawl_date)
                    press_list.append(press)
                    
            except :
                print("해당 섹션의 뉴스 기사가 존재하지 않음.")
                pass
            
        else :
            press_ranking_boxes = soup.select('div.press_ranking_home > div.press_ranking_box')
    
            for box in press_ranking_boxes :
                rankingnews_lists = box.select('ul > li')

                try :
                    for li in rankingnews_lists :
                        ranking = li.select_one('em.list_ranking_num').text
                        ranking_list.append(ranking)

                        title = li.select_one('.list_content > strong').text
                        title_list.append(title)

                        if ranking_type == 'popular' :
                            try :
                                cnt = li.select_one('.list_content > span').text
                                cnt = re.sub("조회수|,", "", cnt)
                                cnt = cnt.strip()
                                num_list.append(cnt)
                            except :
                                num_list.append('미제공')

                        elif ranking_type == 'comment' :
                            try :
                                cnt = li.select_one('.list_content > span').text
                                cnt = re.sub("댓글|,", "", cnt)
                                cnt = cnt.strip()
                                num_list.append(cnt)
                            except :
                                num_list.append('미제공')


                        link = li.select_one('a')["href"]
                        link_list.append(link)

                        date_list.append(crawl_date)
                        press_list.append(press)

                except :
                        continue

    return date_list, press_list, ranking_list, num_list, title_list, link_list


def get_article_contents(link_list) :

    headers = {
        'referer' : 'https://www.naver.com/',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/99.0.4844.51 Safari/537.36'
        }

    content_list, keyword_list = [], []

    for article_link in link_list :
        if article_link == '' :
            content_list.append('')
            keyword_list.append('')
            continue

        res = requests.get(article_link, headers=headers)
        soup = BeautifulSoup(res.text, 'lxml')

        try : 
            content = soup.find('div',{'id' : 'articleBodyContents'}).text # 섹션탭에서 직접 선택하는 경우
        except :
            try : 
                content = soup.find('div', {'id' : 'dic_area'}).text #랭킹뉴스 통하는 경우
            except :
                try :
                    content = soup.find('div', {'id' : 'newsEndContents'}).text # 스포츠
                except :
                    try :
                        content = soup.find('div',{'id' : 'articeBody'}).text # 연예
                    except :
                        continue
        finally :
            content = re.sub('\xa0|\t|\r|\n|', '', content)
            special = re.compile(r'[^ A-Za-z0-9가-힣+]')
            content = special.sub(' ', content)
            content_list.append(content)

            textrank = TextRank(content)
            keyword_list.append(textrank.keywords())
                    
    return content_list, keyword_list


##############
# 키워드 검색 #
##############


def date_setting(keyword, year, month, start_day, end_day, save_path):

    year = int(year)
    month = int(month)
    start_day = int(start_day)
    end_day = int(end_day)
    
    for day in tqdm(range(start_day, end_day+1)):
        date_time_obj = datetime(year=year, month=month, day=day)
        target_date = date_time_obj.strftime("%Y%m%d")
        ds_de = date_time_obj.strftime("%Y.%m.%d")

        get_keywordNews(keyword=keyword, save_path=f"{save_path}/{keyword}/{target_date}_{keyword}", target_date=target_date, ds_de=ds_de)


def get_keywordNews(keyword, save_path, target_date, ds_de, sort=0) :

    crawl_date = f"{target_date[:4]}.{target_date[4:6]}.{target_date[6:]}"
    driver = set_chrome_driver()

    encoded_keyword = urllib.parse.quote(keyword)
    url = f"https://search.naver.com/search.naver?where=news&query={encoded_keyword}&sm=tab_opt&sort={sort}&photo=0&field=0&pd=3&ds={ds_de}&de={ds_de}&docid=&related=0&mynews=0&office_type=0&office_section_code=0&news_office_checked=&nso=so%3Ar%2Cp%3Afrom{target_date}to{target_date}&is_sug_officeid=0"
    more_news_base_url = "https://search.naver.com/search.naver"
    driver.get(url)

    date_list, press_list, num_list, title_list, link_list, content_list, keyword_list, more_news_url_list = [], [], [], [], [], [], [], []
    date_list, press_list, title_list, link_list, more_news_url_list = get_keywordNews_infos(driver=driver,
                                                                                            crawl_date=crawl_date,
                                                                                            date_list=date_list,
                                                                                            press_list=press_list,
                                                                                            title_list=title_list,
                                                                                            link_list=link_list,                                                                                    
                                                                                            more_news_base_url=more_news_base_url,
                                                                                            more_news=True)
    
    content_list, keyword_list, num_list, section_list = get_kN_article_content(driver, link_list)
    driver.close()

    if len(more_news_url_list) > 0:
        more_news_url_list = list(set(more_news_url_list))
        for more_news_url in more_news_url_list:
            driver = set_chrome_driver()
            driver.get(more_news_url)
            
            date_list, press_list, title_list, link_list, more_news_url_list = get_keywordNews_infos(driver=driver,
                                                                                        crawl_date=crawl_date,
                                                                                        date_list=date_list,
                                                                                        press_list=press_list,
                                                                                        title_list=title_list,
                                                                                        link_list=link_list)
                                                                                        

            content_list, keyword_list, num_list, section_list = get_kN_article_content(driver, link_list)       
            driver.close()

    ranking_list = [] # 빈 리스트
    data_list = [date_list, press_list, section_list, title_list, link_list, num_list, content_list, keyword_list, ranking_list]    
    ranking_type = 'keyword'
    
    #to_Excel(save_path, ranking_type, data_list)
    #to_Json(save_path, ranking_type, data_list)
    #clustering(save_path, data_list, ranking_type)


def get_keywordNews_infos(driver, crawl_date, date_list, press_list, title_list, link_list, more_news_base_url=None, more_news=False) :
    
    more_news_url_list = []
    while True :
        html_src = driver.page_source
        soup = BeautifulSoup(html_src, 'lxml')

        # 관련뉴스
        more_news_infos = soup.select('a.news_more')
        
        if more_news :
            for more_news_info in more_news_infos:
                more_news_url = f"{more_news_base_url}{more_news_info.get('href')}"
                more_news_url_list.append(more_news_url)
                          
        article_infos = soup.select('div.news_area')
        if not article_infos :
            break

        for article_info in article_infos :
                
            naver_article = article_info.select_one('div.info_group > a:nth-child(3)')
            if not naver_article :
                continue
            else :
                article_link = naver_article['href']
                
            link_list.append(article_link)
            
            # 언론사명
            press_info = article_info.select_one('div.info_group > a.info.press')
            if press_info is None :
                press_info = article_info.select_one('div.info_group > span.info.press')
                
            press = press_info.text.replace('언론사 선정', '')
            press_list.append(press)

            # 기사 제목
            article = article_info.select_one('a.news_tit')
            title = article.get('title')
            title_list.append(title)
            
            # 날짜
            date_list.append(crawl_date)
        
        time.sleep(2.0)

        next_btn_status = soup.select_one('a.btn_next').get('aria-disabled')
        if next_btn_status == 'true' :
            break

        time.sleep(1.0)
        next_page_btn = driver.find_element(By.CSS_SELECTOR, 'a.btn_next').click()
                      
    return date_list, press_list, title_list, link_list, more_news_url_list


def get_kN_article_content(driver, link_list) :
    
    headers = {
        'referer' : 'https://www.naver.com/',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.88 Safari/537.36'
        }
    
    content_list, keyword_list, num_list, section_list = [], [], [], []

    for article_link in link_list :
        if article_link == '' :
            content_list.append('')
            keyword_list.append('')
            num_list.appned('')
            continue

        res = requests.get(article_link, headers=headers)
        soup = BeautifulSoup(res.text, 'lxml')
        time.sleep(3)

        try : 
            content = soup.find('div',{'id' : 'articleBodyContents'}).text
        except :
            try : 
                content = soup.find('div', {'id' : 'dic_area'}).text
            except :
                try :
                    content = soup.find('div', {'id' : 'newsEndContents'}).text
                except :
                    try :
                        content = soup.find('div',{'id' : 'articeBody'}).text
                    except :
                        continue
        finally :
            content = re.sub('\xa0|\t|\r|\n|', '', content)
            special = re.compile(r'[^ A-Za-z0-9가-힣+]')
            content = special.sub(' ', content)
            content_list.append(content)

            textrank = TextRank(content)
            keyword_list.append(textrank.keywords())

        driver = set_chrome_driver()
        driver.get(article_link)
        time.sleep(3)

        html_src = driver.page_source
        soup = BeautifulSoup(html_src, 'lxml')

        try :
            section = soup.select_one('div.media_end_categorize > a > em').text
        except :
            try :
                section = soup.select_one('div.end_ct > div > div.guide_categorization > a > em').text.replace(" 섹션", "")
            except :
                section = '스포츠'
        finally :
            section_list.append(section)

        try :
            cnt = soup.select_one('.media_end_head_info_variety_likeit > ._reactionModule.u_likeit.nv_notrans > a > span.u_likeit_text._count.num').text
        except :
            try :
                cnt = soup.select_one('.btn_wrap > ._reactionModule.u_likeit > a > span.u_likeit_text._count').text  # 연예
            except :
                try :
                    cnt = soup.select_one('.count > ._reactionModule.u_likeit > a > span.u_likeit_text._count').text  # 스포츠
                except :
                    cnt = ''
        finally :
            if cnt == '공감' :
                cnt = '0'
            cnt = cnt.replace(",", "")

            num_list.append(cnt)
            time.sleep(1)
            driver.close()

    return content_list, keyword_list, num_list, section_list

def clustering(save_path, data_list, ranking_type) :

    article_list = []
    for idx in range (0, len(data_list[0])) :
        try :
            article = dict()
            if ranking_type == 'popular' :
                date = data_list[0][idx]
                press = data_list[1][idx]
                ranking = data_list[2][idx]
                views = data_list[3][idx]
                title = data_list[4][idx]
                link = data_list[5][idx]
                content = data_list[6][idx]
                keywords = data_list[7][idx]
                
                article['views'] = views
                article['ranking'] = ranking
                
            elif ranking_type == 'comment' :
                date = data_list[0][idx]
                press = data_list[1][idx]
                ranking = data_list[2][idx]
                comments = int(data_list[3][idx])
                title = data_list[4][idx]
                link = data_list[5][idx]
                content = data_list[6][idx]
                keywords = data_list[7][idx]
                
                article['comments'] = comments
                article['ranking'] = ranking

            elif ranking_type == 'section' :
                date = data_list[0][idx]
                press = data_list[1][idx]
                section = data_list[2][idx]
                ranking = data_list[3][idx]
                views = data_list[4][idx]
                title = data_list[5][idx]
                link = data_list[6][idx]
                content = data_list[7][idx]
                keywords = data_list[8][idx]
                
                article['section'] = section
                article['ranking'] = ranking
                article['views'] = views
                
            article['date'] = date
            article['press'] = press
            article['title'] = title
            article['content'] = content
            article['link'] = link
            article['keywords'] = keywords
 
        except :
            continue
        article_list.append(article)
        
    if ranking_type == 'popular' :
        df = pd.DataFrame(article_list, columns = ['date', 'press', 'ranking', 'views', 'title', 'link', 'content', 'keywords'])

    elif ranking_type == 'comment' :
        df = pd.DataFrame(article_list, columns = ['date', 'press', 'ranking', 'comments', 'title', 'link', 'content', 'keywords'])

    elif ranking_type == 'section' :
        df = pd.DataFrame(article_list, columns = ['date', 'press', 'section', 'ranking', 'views', 'title', 'link', 'content', 'keywords'])

    okt = Okt()
    # 명사 추출
    noun_list = []
    for content in tqdm(df['content']) :
        if content != '' :
            nouns = okt.nouns(str(content)) # 명사만 추출
            noun_list.append(nouns)

    df['nouns'] = noun_list
    
    # 추출한 명사 없으면 제거 (본문 없으면 제거됨)
    drop_index_list = []
    for i, row in df.iterrows() :
        temp_nouns = row['nouns']
        if len(temp_nouns) == 0 :
            drop_index_list.append(i)
    df = df.drop(drop_index_list)

    df.index = range(len(df))

    text = [' '.join(noun) for noun in df['nouns']]

    # min_df = 0.01 : 문서에 1% 미만으로 나타나는 단어 무시
    # ngram_range=(1,5) : 단어 묶음 1개부터 5개까지
    tfidf_vectorizer = TfidfVectorizer(min_df = 0.01, ngram_range=(1,5))
    vector = tfidf_vectorizer.fit_transform(text)
    # 벡터 리스트 형태로 리턴
    vector = vector.toarray()

    # k, random_state 설정
    article_num = len(data_list[0])
    k = int(math.sqrt((article_num/2)/2))
    kmeans = KMeans(n_clusters = k, random_state = 10)

    kmeans.fit(vector)

    # 클러스터 레이블 저장
    cluster_label = kmeans.fit_predict(vector)
    df['cluster_label'] = cluster_label

    # 중복 기사 제거
    # 카테고리 별로 클러스터링 된 기사들 인덱스 중 하나 선택해 비교 기준으로 삼음
    # 비교 기준 기사 제외한 다른 기사들과의 유사도 측정
    # 유사도 0.9 이상이면 중복 기사라 판단 (일단 몇 % 나오는지 보고 조정 필요)

    for n in range(k) :
        print()
        globals()['cluster{}'.format(n)] = df[df['cluster_label'] == n].index
        #print("카테고리 " + str(n) + "에 포함된 기사 인덱스\n", globals()['cluster{}'.format(n)])
        #print()
        comparison_article = df.iloc[globals()['cluster{}'.format(n)][0]]['title']
        #print("유사도 비교 기준 기사 : ", comparison_article)
        #print()
        #print("----------------------------------------------------------------------------------------------------")

        similarity = cosine_similarity(vector[[globals()['cluster{}'.format(n)][0]]], vector[globals()['cluster{}'.format(n)]])
        similarity = similarity.tolist()

        drop_index_list = []
        for idx in range(len(similarity)-1) :
            if (similarity[0][idx+1] >= 0.9) :
                drop_index_list.append(idx)
            else :
                continue
                             
    df = df.drop(drop_index_list)
    df = df.sort_values(by=['cluster_label'])
    df.index = range(len(df))

    if ranking_type == 'popular' :
        final_df = df[['cluster_label', 'date', 'press', 'ranking', 'views', 'title', 'link', 'content', 'keywords']]

    elif ranking_type == 'comment' :
        df = df.sort_values(by=['cluster_label', 'comments'], ascending=[True, False]).groupby('cluster_label').head(10/k)
        if len(df) > 10 :
            num = len(df) - 10
            index = df.sort_values(by='comments', ascending=True).head(num).index
            print(index)
            df = df.drop(index)
        
        df = df.sort_values(by=['comments'], ascending=False)
        df.index = range(len(df))
        final_df = df[['cluster_label', 'date', 'press', 'ranking', 'comments', 'title', 'link', 'content', 'keywords']]

    elif ranking_type == 'section' :
        final_df = df[['cluster_label', 'date', 'press', 'section', 'ranking', 'views', 'title', 'link', 'content', 'keywords']]
 
    save_path = save_path + "_clustering"

    # excel 파일 생성
    if os.path.exists(f"{save_path}.xlsx") :
        os.remove(f"{save_path}.xlsx")
    final_df.to_excel(f"{save_path}.xlsx", index=False)

    # json 파일 생성
    wb = load_workbook(filename=f"{save_path}.xlsx", data_only=True)
    ws = wb.active

    key_list = []
    for col_num in range(1, ws.max_column+1) :
        key_list.append(ws.cell(row=1, column=col_num).value)

    data_dict = {}
    for row_num in range(2, ws.max_row+1) :
        tmp_dict = {}
        for col_num in range(1, ws.max_column+1) :
            val = ws.cell(row=row_num, column=col_num).value
            tmp_dict[key_list[col_num-1]] = val

        data_dict[row_num-1] = tmp_dict

    wb.close()

    j = json.dumps(data_dict, ensure_ascii=False, indent="\t") 
    with open(f"{save_path}.json", 'w', encoding="utf-8") as f:
        f.write(j)


if __name__ == "__main__" :

    print("1. 언론사별 랭킹뉴스\n2. 키워드 검색\n")
    option = input("option (번호 입력) : ")
    save_path = "./crawling_result"

    if option == "1" :
        try :
            os.makedirs(f"{save_path}/언론사별 랭킹뉴스")
        except :
            pass

        target_date = input("검색일 : ")
        ranking_type = input("랭킹 타입 (1. 많이 본 , 2. 댓글 많은) : ")
        if '1' in ranking_type :
            ranking_type = 'popular'
        elif '2' in ranking_type :
            ranking_type = 'comment'
        elif '3' in ranking_type :
            ranking_type = 'section'

        get_rankingNews(save_path=f"{save_path}/언론사별 랭킹뉴스/{target_date}_언론사별 랭킹뉴스", target_date=target_date, ranking_type=ranking_type)

        
    elif option == "2" :
        keyword = input("검색 키워드 : ")
        start_date = input("검색 시작일 : ")
        end_date = input("검색 종료일 : ")
        
        year = start_date[:4]
        month = start_date[4:6]
        start_day = start_date[6:]
        end_day = end_date[6:]

        try : 
            os.makedirs(f"{save_path}/{keyword}")
        except :
            pass

        date_setting(keyword=keyword, year=year, month=month, start_day=start_day, end_day=end_day, save_path=save_path)